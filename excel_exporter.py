import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from decimal import Decimal
from typing import Dict, List
from financial_core import FinancialModel


class ExcelExporter:
    """Excel格式导出器"""
    
    def __init__(self, model: FinancialModel):
        self.model = model
        self.period = model.period
    
    def export_to_excel(self, output_path: str, template_path: str = None):
        """导出财务模型到Excel文件"""
        try:
            # 如果有模板文件，复制模板
            if template_path:
                wb = openpyxl.load_workbook(template_path)
                ws = wb.active
            else:
                # 创建新的工作簿
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "财务分析结果"
            
            # 创建各个工作表
            self._create_cash_flow_sheet(wb)
            self._create_profit_sheet(wb)
            self._create_investment_sheet(wb)
            self._create_summary_sheet(wb)
            
            # 保存文件
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"Excel导出错误: {e}")
            return False
    
    def _create_cash_flow_sheet(self, wb):
        """创建现金流量表工作表"""
        ws = wb.create_sheet("项目投资现金流量表")
        
        # 设置标题样式
        title_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Arial', size=10, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置对齐
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        results = self.model.results
        
        # 写入标题
        ws['A1'] = "项目投资现金流量表"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = center_alignment
        ws.merge_cells('A1:E1')
        
        # 写入表头
        headers = ['年份', '现金流入', '现金流出', '净现金流量', '累计净现金流量']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_alignment
        
        # 写入数据
        for year_idx in range(self.period.total_period):
            year = year_idx + 1
            row = year_idx + 3
            
            ws.cell(row=row, column=1).value = year
            ws.cell(row=row, column=2).value = float(results.annual_cash_flow_in[year_idx])
            ws.cell(row=row, column=3).value = float(results.annual_cash_flow_out[year_idx])
            ws.cell(row=row, column=4).value = float(results.annual_net_cash_flow[year_idx])
            ws.cell(row=row, column=5).value = float(results.cumulative_cash_flow[year_idx])
            
            # 设置边框和对齐
            for col_idx in range(1, 6):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_alignment
        
        # 调整列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
    
    def _create_profit_sheet(self, wb):
        """创建利润表工作表"""
        ws = wb.create_sheet("利润表")
        
        # 设置样式
        title_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Arial', size=10, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        results = self.model.results
        
        # 写入标题
        ws['A1'] = "利润表"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = center_alignment
        ws.merge_cells('A1:H1')
        
        # 写入表头
        headers = ['年份', '营业收入', '营业成本', '折旧', '摊销', 
                   '税前利润', '所得税', '税后利润']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_alignment
        
        # 写入数据
        for year_idx in range(self.period.total_period):
            year = year_idx + 1
            row = year_idx + 3
            
            ws.cell(row=row, column=1).value = year
            ws.cell(row=row, column=2).value = float(results.annual_revenue[year_idx])
            ws.cell(row=row, column=3).value = float(results.annual_cost[year_idx])
            ws.cell(row=row, column=4).value = float(results.annual_depreciation[year_idx])
            ws.cell(row=row, column=5).value = float(results.annual_amortization[year_idx])
            ws.cell(row=row, column=6).value = float(results.annual_profit_before_tax[year_idx])
            ws.cell(row=row, column=7).value = float(results.annual_income_tax[year_idx])
            ws.cell(row=row, column=8).value = float(results.annual_profit_after_tax[year_idx])
            
            # 设置边框和对齐
            for col_idx in range(1, 9):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_alignment
        
        # 调整列宽
        for col_idx in range(1, 9):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18
    
    def _create_investment_sheet(self, wb):
        """创建投资表工作表"""
        ws = wb.create_sheet("投资汇总")
        
        # 设置样式
        title_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Arial', size=10, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入标题
        ws['A1'] = "投资汇总表"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = center_alignment
        ws.merge_cells('A1:B1')
        
        # 写入数据
        investment_data = [
            ['项目', '金额（万元）'],
            ['工程费', float(self.model.investment.building_cost + 
                            self.model.investment.equipment_procurement_cost + 
                            self.model.investment.equipment_installation_cost)],
            ['工程建设其他费', float(self.model.investment.construction_management_fee + 
                                  self.model.investment.technical_consulting_fee + 
                                  self.model.investment.infrastructure_fee + 
                                  self.model.investment.land_use_fee + 
                                  self.model.investment.patent_fee + 
                                  self.model.investment.other_preparation_fee)],
            ['预备费', float(self.model.investment.basic_contingency_reserve + 
                           self.model.investment.price_contingency_reserve)],
            ['建设期利息', float(self.model.investment.construction_interest)],
            ['流动资金', float(self.model.investment.working_capital)],
            ['项目总投资', float(self.model.results.total_investment)]
        ]
        
        for row_idx, row_data in enumerate(investment_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx + 1, column=col_idx)
                cell.value = value
                cell.border = thin_border
                cell.alignment = center_alignment
                
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_summary_sheet(self, wb):
        """创建财务指标汇总表工作表"""
        ws = wb.create_sheet("财务指标汇总")
        
        # 设置样式
        title_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Arial', size=10, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        results = self.model.results
        
        # 写入标题
        ws['A1'] = "财务指标汇总表"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = center_alignment
        ws.merge_cells('A1:B1')
        
        # 写入数据
        summary_data = [
            ['指标', '数值'],
            ['净现值(NPV)', f"{float(results.npv):,.2f}万元"],
            ['内部收益率(IRR)', f"{float(results.irr):.2%}"],
            ['静态投资回收期', f"{results.static_payback_period:.2f}年" if results.static_payback_period else "未回收"],
            ['动态投资回收期', f"{results.dynamic_payback_period:.2f}年" if results.dynamic_payback_period else "未回收"],
            ['项目总投资', f"{float(results.total_investment):,.2f}万元"],
            ['建设期', f"{self.period.construction_period}年"],
            ['运营期', f"{self.period.operation_period}年"],
            ['计算期', f"{self.period.total_period}年"]
        ]
        
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx + 1, column=col_idx)
                cell.value = value
                cell.border = thin_border
                cell.alignment = center_alignment
                
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30